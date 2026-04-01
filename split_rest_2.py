"""
分割休息②
作業明細→分割開始/終了1-2の転記
①の後に実行する。休息採用が空でない行にのみ転記。
"""

import csv
import os
from datetime import datetime, timedelta, time
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from split_rest_1 import _try_to_date, _find_col, _ensure_len


def _normalize_run_no(v: Any) -> str:
    """運行番号を照合用に正規化（101.0 → 101 など）"""
    if v is None or v == "":
        return ""
    if isinstance(v, (int, float)):
        return str(int(v)).strip()
    s = str(v).strip()
    try:
        return str(int(float(s)))
    except (TypeError, ValueError):
        return s


def _get_time_from_cell(v: Any) -> Optional[time]:
    """時刻として解釈（datetime/time/Excelシリアル/文字列 H:MM:SS, H:MM など）"""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.time()
    if isinstance(v, time):
        return v
    try:
        n = float(v)
        frac = n - int(n)
        if frac < 0:
            frac += 1
        sec = int(frac * 86400)
        return time(sec // 3600, (sec % 3600) // 60, sec % 60)
    except (TypeError, ValueError):
        pass
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        for fmt in ("%H:%M:%S", "%H:%M", "%H:%M:%S.%f"):
            try:
                dt = datetime.strptime(s, fmt)
                return dt.time()
            except ValueError:
                continue
    return None


def _read_work_detail(path: str) -> Dict[str, Tuple[Optional[datetime], Optional[datetime], Optional[datetime], Optional[datetime]]]:
    """
    作業明細から dict[運行番号] = (開始1, 終了1, 開始2, 終了2) を構築
    最大2セット、3回目以降は無視
    """
    result: Dict[str, List] = {}

    def _ensure_rec(key: str):
        if key not in result:
            result[key] = [None, None, None, None, 0]

    _, ext = os.path.splitext(path)
    ext = ext.lower()

    if ext in [".xlsx", ".xlsm"]:
        wb = load_workbook(path, data_only=True)
        ws = wb.worksheets[0]
        max_row = ws.max_row or 0
        if max_row < 2:
            return {}

        header = [ws.cell(1, c).value for c in range(1, 30)]
        i_no = _find_col(header, "運行番号") + 1
        i_date = _find_col(header, "運行日") + 1
        i_arr = _find_col(header, "到着時刻") + 1
        i_dep = _find_col(header, "出発時刻") + 1
        if i_no <= 0 or i_date <= 0 or i_arr <= 0 or i_dep <= 0:
            return {}

        for r in range(2, max_row + 1):
            run_no = _normalize_run_no(ws.cell(r, i_no).value)
            if not run_no:
                continue

            d = _try_to_date(ws.cell(r, i_date).value)
            if not d:
                continue

            t_arr = _get_time_from_cell(ws.cell(r, i_arr).value)
            t_dep = _get_time_from_cell(ws.cell(r, i_dep).value)
            if not t_arr and not t_dep:
                continue

            base_date = d.date()
            arr_dt = None
            dep_dt = None
            if t_arr and t_dep:
                arr_dt = datetime.combine(base_date, t_arr)
                dep_dt = datetime.combine(base_date, t_dep)
                if arr_dt > dep_dt:
                    dep_dt += timedelta(days=1)
            elif t_arr:
                arr_dt = datetime.combine(base_date, t_arr)
            elif t_dep:
                dep_dt = datetime.combine(base_date, t_dep)

            key = run_no
            _ensure_rec(key)
            rec = result[key]
            cnt = rec[4]
            if cnt == 0:
                rec[0], rec[1], rec[4] = arr_dt, dep_dt, 1
            elif cnt == 1:
                rec[2], rec[3], rec[4] = arr_dt, dep_dt, 2
            result[key] = rec

    elif ext == ".csv":
        try:
            with open(path, "r", encoding="cp932", newline="") as f:
                raw = list(csv.reader(f))
        except UnicodeDecodeError:
            with open(path, "r", encoding="utf-8", newline="") as f:
                raw = list(csv.reader(f))
        if len(raw) < 2:
            return {}

        header = raw[0]
        i_no = _find_col(header, "運行番号")
        i_date = _find_col(header, "運行日")
        i_arr = _find_col(header, "到着時刻")
        i_dep = _find_col(header, "出発時刻")
        if i_no < 0 or i_date < 0 or i_arr < 0 or i_dep < 0:
            return {}

        for row in raw[1:]:
            if max(i_no, i_date, i_arr, i_dep) >= len(row):
                continue
            run_no = _normalize_run_no(row[i_no])
            if not run_no:
                continue

            d = _try_to_date(row[i_date])
            if not d:
                continue

            t_arr = _get_time_from_cell(row[i_arr])
            t_dep = _get_time_from_cell(row[i_dep])
            if not t_arr and not t_dep:
                continue

            base_date = d.date()
            arr_dt = None
            dep_dt = None
            if t_arr and t_dep:
                arr_dt = datetime.combine(base_date, t_arr)
                dep_dt = datetime.combine(base_date, t_dep)
                if arr_dt > dep_dt:
                    dep_dt += timedelta(days=1)
            elif t_arr:
                arr_dt = datetime.combine(base_date, t_arr)
            elif t_dep:
                dep_dt = datetime.combine(base_date, t_dep)

            key = run_no
            _ensure_rec(key)
            rec = result[key]
            cnt = rec[4]
            if cnt == 0:
                rec[0], rec[1], rec[4] = arr_dt, dep_dt, 1
            elif cnt == 1:
                rec[2], rec[3], rec[4] = arr_dt, dep_dt, 2
            result[key] = rec

    else:
        return {}

    return {k: (v[0], v[1], v[2], v[3]) for k, v in result.items()}


def process_split_rest_step2(
    rows: List[List],
    work_detail_path: str,
) -> None:
    """
    分割休息②：作業明細→分割開始/終了1-2の転記
    - 休息採用が空の行は転記しない
    - 出庫点呼日時基準で分割開始 < 出庫点呼日時 なら +1日補正
    """
    if len(rows) < 2:
        return

    header = rows[0]
    idx_no = _find_col(header, "運行番号")
    idx_adopt = _find_col(header, "休息採用")
    idx_dep_tenko = _find_col(header, "出庫点呼日時")
    idx_start1 = _find_col(header, "分割開始1")
    idx_end1 = _find_col(header, "分割終了1")
    idx_start2 = _find_col(header, "分割開始2")
    idx_end2 = _find_col(header, "分割終了2")

    if idx_no < 0 or idx_adopt < 0 or idx_dep_tenko < 0 or idx_start1 < 0 or idx_end1 < 0 or idx_start2 < 0 or idx_end2 < 0:
        return

    dict_data = _read_work_detail(work_detail_path)

    for i in range(1, len(rows)):
        row = rows[i]
        _ensure_len(row, max(idx_adopt, idx_dep_tenko, idx_start1, idx_end1, idx_start2, idx_end2) + 1)

        adopt_val = row[idx_adopt] if idx_adopt < len(row) else None
        if adopt_val is None or (isinstance(adopt_val, str) and adopt_val.strip() == ""):
            continue

        key = _normalize_run_no(row[idx_no] if idx_no < len(row) else None)

        # 転記前にクリア
        row[idx_start1] = None
        row[idx_end1] = None
        row[idx_start2] = None
        row[idx_end2] = None

        if not key or key not in dict_data:
            continue

        s1, e1, s2, e2 = dict_data[key]

        row[idx_start1] = s1
        row[idx_end1] = e1
        row[idx_start2] = s2
        row[idx_end2] = e2

        dt_out = _try_to_date(row[idx_dep_tenko])
        if dt_out:
            if s1 and s1 < dt_out:
                row[idx_start1] = s1 + timedelta(days=1)
                if e1:
                    row[idx_end1] = e1 + timedelta(days=1)
            if s2 and s2 < dt_out:
                row[idx_start2] = s2 + timedelta(days=1)
                if e2:
                    row[idx_end2] = e2 + timedelta(days=1)
