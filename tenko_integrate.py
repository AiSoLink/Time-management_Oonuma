"""
点呼データ統合モジュール
対面ファイル・遠隔ファイルを読み込み、統合データを返す。
VBA modTenkoIntegrate / modTenkoNormalize のロジックを踏襲。
"""

import csv
import os
import random
from datetime import datetime, timedelta
from collections import defaultdict
from typing import List, Optional

from openpyxl import load_workbook

# Excelシリアル日付の起点（Windows）
EXCEL_EPOCH = datetime(1899, 12, 30)
# 全角スペース
ZENKAKU_SPACE = "\u3000"


# ============================================================
# 氏名正規化（modTenkoNormalize 踏襲）
# ============================================================

def _compress_spaces(s: str) -> str:
    """連続スペースを1つに圧縮"""
    out = []
    prev_space = False
    for ch in s:
        if ch == " ":
            if not prev_space:
                out.append(" ")
                prev_space = True
        else:
            out.append(ch)
            prev_space = False
    return "".join(out)


def normalize_display_name(s) -> str:
    """
    表示名正規化（VBA NormalizeDisplayName 相当）
    - 全角スペース→半角に統一
    - 前後空白除去
    - 連続スペースを1つに圧縮
    - 空白で分割した部分を全角スペース1個で結合
    - 空白が無い場合はそのまま
    """
    if s is None:
        return ""
    s = str(s)
    s = s.replace(ZENKAKU_SPACE, " ")
    s = s.strip()
    s = _compress_spaces(s)

    if " " not in s:
        return s

    parts = [p for p in s.split(" ") if p]
    return ZENKAKU_SPACE.join(parts)


# ============================================================
# 日時解釈（TryGetDateTime 相当）
# ============================================================

def try_get_datetime(v) -> Optional[datetime]:
    """
    日時として解釈可能なら datetime を返す。
    - 日付型・日付文字列 → 日時化
    - Excelシリアル値（数値）→ 1899/12/30 起点で日時化
    - どれも不可 → None
    """
    if v is None or v == "":
        return None

    # すでに datetime
    if isinstance(v, datetime):
        return v

    # 日付型（date のみ）
    try:
        from datetime import date
        if isinstance(v, date) and not isinstance(v, datetime):
            return datetime.combine(v, datetime.min.time())
    except ImportError:
        pass

    # 日付文字列・Excel互換の解釈
    if isinstance(v, str):
        v = v.strip()
        if not v:
            return None
        # よくある形式を試す
        for fmt in (
            "%Y/%m/%d %H:%M:%S.%f", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M",
            "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
            "%Y/%m/%d", "%Y-%m-%d",
        ):
            try:
                return datetime.strptime(v, fmt)
            except ValueError:
                continue

    # Excelシリアル値
    try:
        n = float(v)
        if n > 0:
            return EXCEL_EPOCH + timedelta(days=n)
    except (TypeError, ValueError):
        pass

    return None


# ============================================================
# 対面：入出庫→点呼区分（MapFaceSectionOrUnselected 相当）
# ============================================================

def _map_face_section(io_val: str) -> str:
    """入出庫文字列から点呼区分を判定"""
    if io_val is None:
        return "未選択"
    s = str(io_val).strip()
    if "出庫" in s or "出車" in s:
        return "出庫"
    if "入庫" in s or "帰庫" in s:
        return "帰庫"
    return "未選択"


# ============================================================
# ファイル読み込み
# ============================================================

def _read_file_to_rows(path: str) -> list[list]:
    """
    ファイル（Excel/CSV）を読み込み、行のリストを返す。
    各行は列のリスト。不足列は None でパディング（最大27列＝A～AA）
    """
    _, ext = os.path.splitext(path)
    ext = ext.lower()
    min_cols = 27  # AA列まで対応

    if ext in [".xlsx", ".xlsm"]:
        wb = load_workbook(path, data_only=True)
        ws = wb.worksheets[0]
        max_row = ws.max_row or 0
        max_col = max(ws.max_column or 0, min_cols)
        rows = []
        for r in range(1, max_row + 1):
            row = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            row = row + [None] * max(0, min_cols - len(row))
            rows.append(row[:min_cols])
        return rows

    if ext == ".csv":
        try:
            with open(path, "r", encoding="cp932", newline="") as f:
                reader = csv.reader(f)
                rows = list(reader)
        except UnicodeDecodeError:
            with open(path, "r", encoding="utf-8", newline="") as f:
                reader = csv.reader(f)
                rows = list(reader)

        result = []
        for row in rows:
            padded = list(row) + [None] * max(0, min_cols - len(row))
            result.append(padded[:min_cols])
        return result

    return []


def _get_cell(row: list, col_idx: int):
    """行から列を取得。インデックスオーバーなら None"""
    if col_idx >= len(row):
        return None
    return row[col_idx]


def _to_emp_id_value(v) -> object:
    """社員番号を数値に変換。不可能な場合は文字列のまま"""
    if v is None or v == "":
        return ""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return int(v) if v == int(v) else v
    s = str(v).strip()
    if not s:
        return ""
    try:
        n = float(s)
        return int(n) if n == int(n) else n
    except ValueError:
        return s


# ============================================================
# メイン：統合データ作成
# ============================================================

def integrate_tenko_data(face_path: str, remote_path: str) -> List[List]:
    """
    対面・遠隔ファイルを読み込み、統合データを返す。

    Args:
        face_path: 対面ファイルのパス
        remote_path: 遠隔ファイルのパス

    Returns:
        [ヘッダー行, データ行1, データ行2, ...]
        ヘッダー: ["社員番号", "社員名", "点呼日時", "点呼区分", "点呼方法"]
        データ行: [社員番号, 社員名, datetime, 点呼区分, 点呼方法]
    """
    header = ["社員番号", "社員名", "点呼日時", "点呼区分", "点呼方法"]
    out_rows = []

    # 対面ファイル（A=0, B=1, C=2, D=3）
    face_rows = _read_file_to_rows(face_path)
    for row in face_rows[1:]:  # ヘッダー行をスキップ
        name_raw = _get_cell(row, 1)
        if not name_raw or str(name_raw).strip() == "":
            continue

        dt = try_get_datetime(_get_cell(row, 3))
        if dt is None:
            continue

        emp_id = _to_emp_id_value(_get_cell(row, 0))
        io_raw = _get_cell(row, 2)
        io_raw = "" if io_raw is None else str(io_raw)
        section = _map_face_section(io_raw)
        name_disp = normalize_display_name(name_raw)
        out_rows.append([emp_id, name_disp, dt, section, "対面"])

    # 遠隔ファイル（A=0, B=1, E=4, P=15, AA=26）
    remote_rows = _read_file_to_rows(remote_path)
    for row in remote_rows[1:]:
        name_raw = _get_cell(row, 1)
        if not name_raw or str(name_raw).strip() == "":
            continue

        emp_id = _to_emp_id_value(_get_cell(row, 0))
        name_disp = normalize_display_name(name_raw)

        for col_idx, section in [(4, "出庫"), (15, "中間"), (26, "帰庫")]:
            dt = try_get_datetime(_get_cell(row, col_idx))
            if dt is not None:
                out_rows.append([emp_id, name_disp, dt, section, "電話"])

    # ソート：社員名（index 1）昇順 → 点呼日時（index 2）昇順
    out_rows.sort(key=lambda r: (str(r[1]), r[2]))

    return [header] + out_rows


# ============================================================
# 完成ファイルへの点呼データ転記
# ============================================================

def _normalize_key_for_match(v) -> str:
    """突合用：乗務員コード・社員番号を比較用文字列に正規化"""
    if v is None or v == "":
        return ""
    return str(v).strip()


def _random_datetime_between(start_dt: datetime, end_dt: datetime) -> datetime:
    """開始～終了の範囲で秒単位のランダム日時を返す（両端含む）"""
    if end_dt < start_dt:
        start_dt, end_dt = end_dt, start_dt
    total_sec = int((end_dt - start_dt).total_seconds())
    if total_sec <= 0:
        return start_dt
    return start_dt + timedelta(seconds=random.randint(0, total_sec))


def fill_tenko_into_rows(
    completed_rows: List[List],
    tenko_rows: List[List],
    keep_columns: List[str],
    dep_minutes: float = 60.0,
    ret_minutes: float = 60.0,
) -> set[tuple[int, int]]:
    """
    完成ファイルのU～Z列に点呼データを転記する（completed_rows を in-place で更新）

    - 突合キー: 乗務員コード(C) = 社員番号(A)
    - 出庫: 出庫日時(M) ± dep_minutes 分の範囲内、最速の点呼を採用
    - 帰庫: 帰庫日時(N) ± ret_minutes 分の範囲内、最遅の点呼を採用
    - 中間: 出庫点呼日時(U)～帰庫点呼日時(Y) の間、最初の1件を採用
    Returns:
        fallbackで補完したセル座標の集合（completed_rows基準、0始まり）
    """
    fallback_cells: set[tuple[int, int]] = set()
    try:
        idx_crew = keep_columns.index("乗務員コード")
        idx_dep = keep_columns.index("出庫日時")
        idx_ret = keep_columns.index("帰庫日時")
        idx_dep_tenko = keep_columns.index("出庫点呼日時")
        idx_dep_method = keep_columns.index("出庫点呼方法")
        idx_mid_tenko = keep_columns.index("中間点呼日時")
        idx_mid_method = keep_columns.index("中間点呼方法")
        idx_ret_tenko = keep_columns.index("帰庫点呼日時")
        idx_ret_method = keep_columns.index("帰庫点呼方法")
    except ValueError:
        return fallback_cells

    if len(tenko_rows) < 2:  # ヘッダーのみ
        return fallback_cells

    tenko_data = tenko_rows[1:]  # ヘッダー除外
    # tenko: [社員番号0, 社員名1, 点呼日時2, 点呼区分3, 点呼方法4]

    # 社員番号ごとに点呼データをグループ化
    by_emp = defaultdict(list)
    for row in tenko_data:
        emp_key = _normalize_key_for_match(row[0])
        if emp_key:
            by_emp[emp_key].append(row)

    dep_delta = timedelta(minutes=dep_minutes)
    ret_delta = timedelta(minutes=ret_minutes)
    fallback_delta = timedelta(minutes=10)

    for row_idx, crow in enumerate(completed_rows[1:], start=1):  # ヘッダーをスキップ
        if len(crow) <= max(idx_ret_method, idx_ret_tenko):
            continue
        crew_key = _normalize_key_for_match(crow[idx_crew])
        tenko_list = by_emp.get(crew_key, []) if crew_key else []
        dep_dt = try_get_datetime(crow[idx_dep])
        ret_dt = try_get_datetime(crow[idx_ret])

        # 1) 出庫: 出庫日時±dep_minutes 分の範囲内、最速
        if dep_dt:
            dep_candidates = [
                r for r in tenko_list
                if r[3] == "出庫" and isinstance(r[2], datetime)
                and (dep_dt - dep_delta) <= r[2] <= (dep_dt + dep_delta)
            ]
            if dep_candidates:
                earliest = min(dep_candidates, key=lambda r: r[2])
                crow[idx_dep_tenko] = earliest[2]
                crow[idx_dep_method] = earliest[4]

        # 2) 帰庫: 帰庫日時±ret_minutes 分の範囲内、最遅
        if ret_dt:
            ret_candidates = [
                r for r in tenko_list
                if r[3] == "帰庫" and isinstance(r[2], datetime)
                and (ret_dt - ret_delta) <= r[2] <= (ret_dt + ret_delta)
            ]
            if ret_candidates:
                latest = max(ret_candidates, key=lambda r: r[2])
                crow[idx_ret_tenko] = latest[2]
                crow[idx_ret_method] = latest[4]

        # 2.5) 紐づかなかった場合はランダム補完
        # - U(出庫点呼日時): M(出庫日時)-10分 ～ M の範囲
        # - Y(帰庫点呼日時): N(帰庫日時) ～ N+10分 の範囲
        if dep_dt and not try_get_datetime(crow[idx_dep_tenko]):
            crow[idx_dep_tenko] = _random_datetime_between(dep_dt - fallback_delta, dep_dt)
            fallback_cells.add((row_idx, idx_dep_tenko))
        if ret_dt and not try_get_datetime(crow[idx_ret_tenko]):
            crow[idx_ret_tenko] = _random_datetime_between(ret_dt, ret_dt + fallback_delta)
            fallback_cells.add((row_idx, idx_ret_tenko))

        # 3) 中間: U～Y の間、最初の1件（U,Yは上で埋めた後の値）
        u_val = crow[idx_dep_tenko]
        y_val = crow[idx_ret_tenko]
        u_dt = try_get_datetime(u_val) if u_val else None
        y_dt = try_get_datetime(y_val) if y_val else None
        if u_dt and y_dt and u_dt < y_dt:
            mid_candidates = [
                r for r in tenko_list
                if r[3] == "中間" and isinstance(r[2], datetime)
                and u_dt <= r[2] <= y_dt
            ]
            if mid_candidates:
                first_mid = min(mid_candidates, key=lambda r: r[2])
                crow[idx_mid_tenko] = first_mid[2]
                crow[idx_mid_method] = first_mid[4]

    return fallback_cells
