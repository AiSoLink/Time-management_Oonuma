import json
import os
import sys
import csv
import shutil
import tempfile
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from tenko_integrate import integrate_tenko_data, fill_tenko_into_rows
from monthly_merge import merge_monthly_files
from split_rest import run_split_rest

APP_TITLE = "大沼運輸倉庫_SGシステム → 一番星と点呼記録簿"
OUTPUT_SHEETNAME_DAILY = "毎日ファイル"   # 毎日タブで出力するExcelのシート名
OUTPUT_SHEETNAME = "月末ファイル"         # 月末タブで出力するExcelのシート名

# 元データから抽出する列（ソースに存在する前提）
EXTRACT_COLUMNS = [
    "運行番号",        # A列
    "運行日",          # B列
    "乗務員コード",    # C列
    "乗務員名",        # D列
    "車両コード",      # E列
    "車両番号",        # G列
    "出庫メーター",    # H列
    "帰庫メーター",    # I列
    "休憩",            # J列
    "休息",            # K列
    "出庫日時",        # M列
    "帰庫日時",        # N列
]

# 挿入する列（ソースに存在しない。指定位置に空欄で追加）
INSERT_COLUMNS = [
    "枝番",            # F列
    "休息採用",        # L列
    "運行間休息",      # O列
    "ランキング",      # P列
    "分割開始1",       # Q列
    "分割終了1",       # R列
    "分割開始2",       # S列
    "分割終了2",       # T列
    "出庫点呼日時",    # U列
    "出庫点呼方法",    # V列
    "中間点呼日時",    # W列
    "中間点呼方法",    # X列
    "帰庫点呼日時",    # Y列
    "帰庫点呼方法",    # Z列
]

# 統合後の出力列（この順序で出力）
KEEP_COLUMNS = [
    "運行番号",        # A列
    "運行日",          # B列
    "乗務員コード",    # C列
    "乗務員名",        # D列
    "車両コード",      # E列
    "枝番",            # F列（挿入）
    "車両番号",        # G列
    "出庫メーター",    # H列
    "帰庫メーター",    # I列
    "休憩",            # J列
    "休息",            # K列
    "休息採用",        # L列（挿入）
    "出庫日時",        # M列
    "帰庫日時",        # N列
    "運行間休息",      # O列（挿入）
    "ランキング",      # P列（挿入）
    "分割開始1",       # Q列（挿入）
    "分割終了1",       # R列（挿入）
    "分割開始2",       # S列（挿入）
    "分割終了2",       # T列（挿入）
    "出庫点呼日時",    # U列（挿入）
    "出庫点呼方法",    # V列（挿入）
    "中間点呼日時",    # W列（挿入）
    "中間点呼方法",    # X列（挿入）
    "帰庫点呼日時",    # Y列（挿入）
    "帰庫点呼方法",    # Z列（挿入）
    "運転",            # AA列（ソースZ列から取得）
]

# フォルダ名の定義
FOLDER_TRASH = "ゴミ箱"
FOLDER_OUTPUT = "完成フォルダ"
FOLDER_OUTPUT_DAILY = "毎日"   # 完成フォルダ内
FOLDER_OUTPUT_MONTHLY = "月末"  # 完成フォルダ内

# 点呼設定（非表示フォルダ内に保存）
FOLDER_CONFIG = "._sg_config"
TENKO_SETTINGS_FILE = "tenko_settings.json"


def _get_tenko_settings_path(base_dir: str) -> str:
    """点呼設定ファイルのフルパス（非表示フォルダ内）"""
    return os.path.join(base_dir, FOLDER_CONFIG, TENKO_SETTINGS_FILE)


def _ensure_hidden_config_dir(base_dir: str) -> str:
    """
    設定用フォルダを作成し、Windowsでは非表示にする。
    非表示フォルダでも読み書きは通常通り可能。
    """
    config_dir = os.path.join(base_dir, FOLDER_CONFIG)
    if not os.path.isdir(config_dir):
        os.makedirs(config_dir, exist_ok=True)
        # Windows: フォルダを非表示にする
        if sys.platform == "win32":
            try:
                import ctypes
                ctypes.windll.kernel32.SetFileAttributesW(config_dir, 0x02)  # FILE_ATTRIBUTE_HIDDEN
            except Exception:
                pass
    return config_dir


def _user_friendly_error_message(exc: Exception) -> str:
    """例外をIT未経験者にも分かる日本語メッセージに変換"""
    err = exc
    errstr = str(err).lower() if err else ""
    if isinstance(err, PermissionError) or "permission" in errstr or "access" in errstr or "denied" in errstr:
        return "同じ名前のファイルが既に開いているか、使用中です。\nそのファイル（Excel等）を閉じてから、再度お試しください。"
    if isinstance(err, FileNotFoundError):
        return "指定したファイルまたはフォルダが見つかりません。\nパスを確認してください。"
    if isinstance(err, OSError) and hasattr(err, "errno"):
        if err.errno == 28:  # ENOSPC
            return "ディスクの空き容量が不足しています。"
        if err.errno == 13:  # EACCES
            return "保存先に書き込み権限がありません。\n別のフォルダを指定するか、管理者権限を確認してください。"
    if "openpyxl" in errstr or "excel" in errstr or "workbook" in errstr:
        return "Excelファイルの読み込みに失敗しました。\nファイルが破損しているか、別のアプリで開いていないか確認してください。"
    if "decode" in errstr or "encoding" in errstr:
        return "ファイルの文字コードが読み取れません。\n別のファイルでお試しください。"
    return f"予期せぬエラーが発生しました。\n\n詳細: {err}"


def get_base_dir():
    """
    exe化された場合と通常実行の場合の両方に対応して、
    アプリケーションのベースディレクトリを取得する
    """
    if getattr(sys, 'frozen', False):
        # exe化されている場合
        return os.path.dirname(sys.executable)
    else:
        # 通常実行の場合
        return os.path.dirname(os.path.abspath(__file__))


def _load_tenko_settings(base_dir: str) -> tuple:
    """
    点呼の±分設定を読み込み。(dep_minutes, ret_minutes) を返す。
    初回起動時は非表示フォルダを作成し、デフォルト設定ファイルを生成する。
    """
    _ensure_hidden_config_dir(base_dir)
    path = _get_tenko_settings_path(base_dir)
    try:
        if os.path.isfile(path):
            with open(path, "r", encoding="utf-8") as f:
                d = json.load(f)
                # 新形式: dep_minutes, ret_minutes / 旧形式: dep_hours, ret_hours を分に変換
                dep = d.get("dep_minutes")
                ret = d.get("ret_minutes")
                if dep is None:
                    dep = float(d.get("dep_hours", 1)) * 60
                else:
                    dep = float(dep)
                if ret is None:
                    ret = float(d.get("ret_hours", 1)) * 60
                else:
                    ret = float(ret)
                return (max(1, min(1440, dep)), max(1, min(1440, ret)))
    except (json.JSONDecodeError, TypeError, ValueError):
        pass
    # ファイルが無い場合はデフォルトで作成
    default_dep, default_ret = 60.0, 60.0
    _save_tenko_settings(base_dir, default_dep, default_ret)
    return (default_dep, default_ret)


def _save_tenko_settings(base_dir: str, dep_minutes: float, ret_minutes: float) -> None:
    """点呼の±分設定を非表示フォルダ内に保存。"""
    _ensure_hidden_config_dir(base_dir)
    path = _get_tenko_settings_path(base_dir)
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump({"dep_minutes": dep_minutes, "ret_minutes": ret_minutes}, f, ensure_ascii=False)
    except (IOError, OSError):
        pass


def _read_file_to_rows(path: str):
    """ファイル（Excel/CSV）を読み込み、行のリストを返す。"""
    _, ext = os.path.splitext(path)
    ext = ext.lower()
    if ext in [".xlsx", ".xlsm"]:
        wb = load_workbook(path, data_only=True)
        ws = wb.worksheets[0]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0:
            return []
        return [[ws.cell(row=r, column=c).value for c in range(1, max_col + 1)] for r in range(1, max_row + 1)]
    if ext == ".csv":
        try:
            with open(path, "r", encoding="cp932", newline="") as f:
                rows = list(csv.reader(f))
        except UnicodeDecodeError:
            with open(path, "r", encoding="utf-8", newline="") as f:
                rows = list(csv.reader(f))
        return [[_convert_cell_value(cell) for cell in row] for row in rows]
    return []


def _merge_work_detail_files(paths: list) -> str:
    """
    複数の作業明細ファイルを統合。先頭ファイルのヘッダーのみ残し、他は削除。
    一時ファイルに保存してパスを返す。
    """
    if not paths:
        return ""
    if len(paths) == 1:
        return paths[0]

    first_rows = _read_file_to_rows(paths[0])
    if not first_rows:
        return paths[0]
    header = list(first_rows[0])
    merged = [header]
    name_to_idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None and str(h).strip()}

    for path_idx, path in enumerate(paths):
        rows = _read_file_to_rows(path)
        if not rows:
            continue
        for row in rows[1:]:
            if not any(cell not in (None, "", " ") for cell in row):
                continue
            if path_idx == 0:
                pad = list(row)
                if len(pad) < len(header):
                    pad.extend([None] * (len(header) - len(pad)))
                merged.append(pad[:len(header)])
            else:
                src_header = rows[0]
                src_idx = {str(src_header[i]).strip(): i for i in range(len(src_header)) if src_header[i] is not None and str(src_header[i]).strip()}
                new_row = [None] * len(header)
                for col_name, dst_i in name_to_idx.items():
                    if col_name in src_idx and src_idx[col_name] < len(row):
                        new_row[dst_i] = row[src_idx[col_name]]
                merged.append(new_row)

    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb = Workbook()
    ws = wb.active
    for row in merged:
        ws.append(row)
    wb.save(tmp_path)
    return tmp_path


def _convert_cell_value(value: str):
    """
    CSV から読み込んだ文字列を、可能であれば数値(int / float)に変換する。
    - 空文字や空白だけの場合はそのまま返す
    - 変換できない場合も元の文字列を返す
    """
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return ""

    # 整数として解釈できるか試す
    try:
        return int(s)
    except ValueError:
        pass

    # 小数として解釈できるか試す
    try:
        return float(s)
    except ValueError:
        pass

    # どちらも無理なら元の文字列
    return value


def _filter_rows_by_columns(all_rows, keep_columns, insert_columns):
    """
    全行データから指定列を抽出し、挿入列を空欄で追加する。
    - 抽出列: ソースの header（1行目）から該当列を検索し、存在すれば値を取得
    - 挿入列: ソースに存在しない前提。常に空欄で出力
    """
    insert_set = set(insert_columns)

    if not all_rows:
        return []

    header = all_rows[0]
    # 列名 → インデックスのマッピング（ソースに存在する列のみ）
    name_to_idx = {}
    for idx, name in enumerate(header):
        n = str(name).strip() if name is not None else ""
        if n and n not in name_to_idx:
            name_to_idx[n] = idx

    # 各 keep_column について: 挿入列なら常に空、そうでなければソースから取得
    def get_cell_value(col_name, row):
        if col_name in insert_set:
            return ""  # 挿入列はソースに存在しない前提で常に空
        i = name_to_idx.get(col_name, -1)
        if i < 0 or i >= len(row):
            return ""
        val = row[i]
        return val if val is not None else ""

    # ヘッダー行は keep_columns で固定
    result = [list(keep_columns)]
    for row in all_rows[1:]:
        new_row = [get_cell_value(col_name, row) for col_name in keep_columns]
        result.append(new_row)

    return result


def _format_untin_display(value):
    """
    運転列の表示用整形。
    - 05:40:47.5 → 5:40:47（先頭0なし・小数点以下なし）
    - ソースがExcel時刻型・文字列・floatのいずれでも対応
    """
    if value is None or (isinstance(value, str) and str(value).strip() == ""):
        return ""
    if hasattr(value, "hour"):
        # datetime.time
        return f"{value.hour}:{value.minute:02d}:{int(value.second)}"
    if isinstance(value, (int, float)):
        # Excelシリアル時刻（1 = 24時間）
        s = float(value) * 86400
        h = int(s // 3600) % 24
        m = int((s % 3600) // 60)
        sec = int(s % 60)
        return f"{h}:{m:02d}:{sec:02d}"
    s = str(value).strip()
    parts = s.split(":")
    if len(parts) >= 3:
        try:
            h = int(parts[0])
            m = int(parts[1])
            sec = int(float(parts[2]))
            return f"{h}:{m:02d}:{sec:02d}"
        except (ValueError, TypeError):
            pass
    return s


def _fill_eda_num_from_vehicle_code(rows, keep_columns):
    """
    車両コード列（2～5桁の数字）の一の桁を切り取り、枝番列に転記する。
    - 車両コード: 一の桁を削除（8290 → 829）
    - 枝番: 切り取った一の桁を数値で保存
    ヘッダー行はスキップし、2行目以降のデータ行のみ処理する。
    """
    try:
        idx_vehicle = keep_columns.index("車両コード")
        idx_eda = keep_columns.index("枝番")
    except ValueError:
        return  # 列が存在しない場合は何もしない

    for row in rows[1:]:  # ヘッダー行をスキップ
        if idx_vehicle >= len(row) or idx_eda >= len(row):
            continue
        val = row[idx_vehicle]
        if val is None or val == "":
            continue
        try:
            num = int(str(val).strip())
            num = abs(num)  # 負の数の場合も考慮
            ones_digit = num % 10
            row[idx_vehicle] = num // 10   # 車両コードから一の桁を削除
            row[idx_eda] = ones_digit      # 枝番に数値で保存
        except (ValueError, TypeError):
            pass  # 数値に変換できない場合はスキップ


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("720x420")
        self.files = []
        self.face_file = None   # 対面アルキラーファイル
        self.remote_file = None  # 遠隔アルキラーファイル
        self.monthly_files = []  # 月末用：毎日の完成ファイル一覧
        self.work_detail_files = []  # 月末用：作業明細ファイル（複数可）
        self.monthly_merged_data = None  # 月末統合結果（後続フローで使用）

        # ベースディレクトリとフォルダパスを設定
        self.base_dir = get_base_dir()
        self.trash_dir = os.path.join(self.base_dir, FOLDER_TRASH)
        self.output_dir = os.path.join(self.base_dir, FOLDER_OUTPUT)
        self.output_dir_daily = os.path.join(self.output_dir, FOLDER_OUTPUT_DAILY)
        self.output_dir_monthly = os.path.join(self.output_dir, FOLDER_OUTPUT_MONTHLY)

        # 点呼の±分設定（設定から読み込み）
        self.tenko_dep_minutes, self.tenko_ret_minutes = _load_tenko_settings(self.base_dir)
        
        # 必要なフォルダを作成（ゴミ箱・完成フォルダ・毎日・月末）
        os.makedirs(self.trash_dir, exist_ok=True)
        os.makedirs(self.output_dir, exist_ok=True)
        os.makedirs(self.output_dir_daily, exist_ok=True)
        os.makedirs(self.output_dir_monthly, exist_ok=True)

        # タブ（毎日用 / 月末用）
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        # 選択中タブを色で強調（clam テーマで色指定を有効化）
        _style = ttk.Style()
        try:
            _style.theme_use("clam")
        except tk.TclError:
            pass
        _style.map("TNotebook.Tab", background=[("selected", "#BBDEFB")], foreground=[("selected", "black")])

        # 毎日用タブ
        self.frame_daily = tk.Frame(self.notebook)
        self.notebook.add(self.frame_daily, text="毎日用")

        # 毎日用：ファイル操作UI
        # 上段：3つの選択ボタン（横並び・幅広）＋ 設定（右上）
        top = tk.Frame(self.frame_daily)
        top.pack(fill="x", padx=10, pady=10)

        tk.Button(top, text="⚙ 設定", command=self.open_tenko_settings, width=8, height=2).pack(side="right", padx=(8, 15))
        btn_width = 24
        tk.Button(top, text="SGシステム運行データ", command=self.add_files, width=btn_width, height=2).pack(side="left", padx=4)
        tk.Button(top, text="対面アルキラー", command=self.select_face_file, width=btn_width, height=2).pack(side="left", padx=4)
        tk.Button(top, text="遠隔アルキラー", command=self.select_remote_file, width=btn_width, height=2).pack(side="left", padx=4)

        mid = tk.Frame(self.frame_daily)
        mid.pack(fill="both", expand=True, padx=10)

        left_d = tk.Frame(mid)
        left_d.pack(side="left", fill="both", expand=True)
        left_d.grid_columnconfigure(0, weight=1)
        left_d.grid_rowconfigure(1, weight=1)
        left_d.grid_rowconfigure(3, weight=1)
        left_d.grid_rowconfigure(5, weight=1)

        tk.Label(left_d, text="運行データ").grid(row=0, column=0, sticky="w")
        list_frame_d = tk.Frame(left_d)
        list_frame_d.grid(row=1, column=0, sticky="nsew")
        list_frame_d.grid_rowconfigure(0, weight=1)
        list_frame_d.grid_columnconfigure(0, weight=1)
        self.listbox = tk.Listbox(list_frame_d, selectmode=tk.EXTENDED)
        self.listbox.grid(row=0, column=0, sticky="nsew")

        tk.Label(left_d, text="対面アルキラー").grid(row=2, column=0, sticky="w", pady=(8, 0))
        face_frame = tk.Frame(left_d)
        face_frame.grid(row=3, column=0, sticky="nsew")
        face_frame.grid_rowconfigure(0, weight=1)
        face_frame.grid_columnconfigure(0, weight=1)
        self.listbox_face = tk.Listbox(face_frame, selectmode=tk.EXTENDED)
        self.listbox_face.grid(row=0, column=0, sticky="nsew")

        tk.Label(left_d, text="遠隔アルキラー").grid(row=4, column=0, sticky="w", pady=(8, 0))
        remote_frame = tk.Frame(left_d)
        remote_frame.grid(row=5, column=0, sticky="nsew")
        remote_frame.grid_rowconfigure(0, weight=1)
        remote_frame.grid_columnconfigure(0, weight=1)
        self.listbox_remote = tk.Listbox(remote_frame, selectmode=tk.EXTENDED)
        self.listbox_remote.grid(row=0, column=0, sticky="nsew")

        # 右側：上下矢印・削除・すべて削除（上から下へ、下寄せ）
        right = tk.Frame(mid)
        right.pack(side="left", fill="y", padx=10, pady=(25, 0))

        btn_side_width = 12
        tk.Button(right, text="▲ 上へ", command=self.move_up, width=btn_side_width, height=1).pack(pady=4)
        tk.Button(right, text="▼ 下へ", command=self.move_down, width=btn_side_width, height=1).pack(pady=4)
        tk.Button(right, text="削除", command=self.remove_selected, width=btn_side_width, height=2).pack(pady=(20, 8))
        tk.Button(right, text="すべて削除", command=self.clear_all, width=btn_side_width, height=2).pack(pady=4)
        btn_merge = tk.Button(right, text="アップロード", command=self.merge_and_export, width=btn_side_width, height=2,
                              bg="#2196F3", fg="white", activebackground="#1976D2", activeforeground="white")
        btn_merge.pack(pady=(50, 0))

        # 月末用タブ（毎日の完成ファイルを1つに統合）
        self.frame_monthly = tk.Frame(self.notebook)
        self.notebook.add(self.frame_monthly, text="月末用")

        mid_m = tk.Frame(self.frame_monthly)
        mid_m.pack(fill="both", expand=True, padx=10, pady=10)

        left_m = tk.Frame(mid_m)
        left_m.pack(side="left", fill="both", expand=True)
        left_m.grid_rowconfigure(2, weight=1)
        left_m.grid_rowconfigure(4, weight=1)
        left_m.grid_columnconfigure(0, weight=1)

        btn_pick_m = tk.Frame(left_m)
        btn_pick_m.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        btn_pick_m.grid_columnconfigure(0, weight=1, uniform="monthly_pick_btns")
        btn_pick_m.grid_columnconfigure(1, weight=1, uniform="monthly_pick_btns")
        tk.Button(btn_pick_m, text="毎日ファイル", command=self.add_monthly_files, height=2).grid(
            row=0, column=0, sticky="ew", padx=(0, 4)
        )
        tk.Button(btn_pick_m, text="SGシステム作業明細", command=self.add_work_detail_files, height=2).grid(
            row=0, column=1, sticky="ew", padx=(4, 0)
        )

        tk.Label(left_m, text="毎日ファイル").grid(row=1, column=0, sticky="w")
        list_frame = tk.Frame(left_m)
        list_frame.grid(row=2, column=0, sticky="nsew")
        list_frame.grid_rowconfigure(0, weight=1)
        list_frame.grid_columnconfigure(0, weight=1)
        self.listbox_monthly = tk.Listbox(list_frame, selectmode=tk.EXTENDED)
        self.listbox_monthly.grid(row=0, column=0, sticky="nsew")

        tk.Label(left_m, text="作業明細").grid(row=3, column=0, sticky="w", pady=(10, 0))
        work_frame = tk.Frame(left_m)
        work_frame.grid(row=4, column=0, sticky="nsew")
        work_frame.grid_rowconfigure(0, weight=1)
        work_frame.grid_columnconfigure(0, weight=1)
        self.listbox_work_detail = tk.Listbox(work_frame, selectmode=tk.EXTENDED)
        self.listbox_work_detail.grid(row=0, column=0, sticky="nsew")

        right_m = tk.Frame(mid_m)
        right_m.pack(side="left", fill="y", padx=10)
        tk.Button(right_m, text="⚙ 設定", command=self.open_tenko_settings, width=12, height=2).pack(pady=(0, 4))
        tk.Label(right_m, text="毎日").pack(anchor="w")
        tk.Button(right_m, text="削除", command=self.remove_selected_monthly, width=12, height=2).pack(pady=4)
        tk.Button(right_m, text="すべて削除", command=self.clear_all_monthly, width=12, height=2).pack(pady=4)
        tk.Label(right_m, text="作業明細").pack(anchor="w", pady=(10, 0))
        tk.Button(right_m, text="削除", command=self.remove_selected_work_detail, width=12, height=2).pack(pady=4)
        tk.Button(right_m, text="すべて削除", command=self.clear_all_work_detail, width=12, height=2).pack(pady=4)
        btn_merge_m = tk.Button(right_m, text="アップロード", command=self.merge_monthly_and_export, width=12, height=2,
                                bg="#2196F3", fg="white", activebackground="#1976D2", activeforeground="white")
        btn_merge_m.pack(pady=(15, 0))

        # 画面中央（横）× 中央と最上の中間（縦）に配置
        self.update_idletasks()
        w, h = 720, 560
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        x = max(0, (sw - w) // 2)
        y = max(0, (sh - h) // 4)  # 最上と中央の中間（1/4位置）
        self.geometry(f"{w}x{h}+{x}+{y}")

    def open_tenko_settings(self):
        """点呼の±分設定ダイアログ"""
        win = tk.Toplevel(self)
        win.title("点呼データ 設定")
        win.geometry("340x200")
        win.transient(self)
        win.grab_set()

        # メインウィンドウの中央に配置
        win.update_idletasks()
        pw, ph = 340, 200
        mx, my = self.winfo_x(), self.winfo_y()
        mw, mh = self.winfo_width(), self.winfo_height()
        x = mx + max(0, (mw - pw) // 2)
        y = my + max(0, (mh - ph) // 2)
        win.geometry(f"{pw}x{ph}+{x}+{y}")

        f = tk.Frame(win, padx=20, pady=20)
        f.pack(fill="both", expand=True)

        tk.Label(f, text="デジタコの出庫・帰庫日時を基準として、\nアルコールのデータは±何分以内？").pack(anchor="w", pady=(0, 10))

        row1 = tk.Frame(f)
        row1.pack(fill="x", pady=4)
        tk.Label(row1, text="出庫日時 ±", width=12, anchor="e").pack(side="left", padx=(0, 8))
        ent_dep = tk.Entry(row1, width=8)
        ent_dep.pack(side="left")
        ent_dep.insert(0, str(int(self.tenko_dep_minutes)))
        tk.Label(row1, text="分").pack(side="left", padx=(4, 0))

        row2 = tk.Frame(f)
        row2.pack(fill="x", pady=4)
        tk.Label(row2, text="帰庫日時 ±", width=12, anchor="e").pack(side="left", padx=(0, 8))
        ent_ret = tk.Entry(row2, width=8)
        ent_ret.pack(side="left")
        ent_ret.insert(0, str(int(self.tenko_ret_minutes)))
        tk.Label(row2, text="分").pack(side="left", padx=(4, 0))

        def save_and_close():
            try:
                dep = float(ent_dep.get().strip() or "60")
                ret = float(ent_ret.get().strip() or "60")
                dep = max(1, min(1440, dep))
                ret = max(1, min(1440, ret))
                self.tenko_dep_minutes = dep
                self.tenko_ret_minutes = ret
                _save_tenko_settings(self.base_dir, dep, ret)
                messagebox.showinfo(APP_TITLE, f"設定を保存しました。\n出庫日時: ±{int(dep)}分\n帰庫日時: ±{int(ret)}分")
                win.destroy()
            except ValueError:
                messagebox.showerror(APP_TITLE, "1～1440 の数値を入力してください。")

        btn_frame = tk.Frame(f)
        btn_frame.pack(fill="x", pady=(16, 0))
        tk.Button(btn_frame, text="保存", command=save_and_close, width=10).pack(side="right", padx=4)
        tk.Button(btn_frame, text="キャンセル", command=win.destroy, width=10).pack(side="right")

    def refresh_list(self):
        self.listbox.delete(0, tk.END)
        for f in self.files:
            # 画面表示はファイル名のみ（フルパスは内部で保持）
            self.listbox.insert(tk.END, os.path.basename(f))

    def refresh_listbox_alkira(self):
        """対面・遠隔アルキラーのリスト表示を更新"""
        self.listbox_face.delete(0, tk.END)
        if self.face_file:
            self.listbox_face.insert(tk.END, os.path.basename(self.face_file))
        self.listbox_remote.delete(0, tk.END)
        if self.remote_file:
            self.listbox_remote.insert(tk.END, os.path.basename(self.remote_file))

    def add_files(self):
        """ファイル選択ダイアログでファイルを追加"""
        paths = filedialog.askopenfilenames(
            title="統合するファイルを選択",
            filetypes=[
                ("Excel / CSV files", "*.xlsx *.xlsm *.csv"),
                ("Excel files", "*.xlsx *.xlsm"),
                ("CSV files", "*.csv"),
                ("All files", "*.*"),
            ]
        )
        if not paths:
            return
        for p in paths:
            if p not in self.files:
                self.files.append(p)
        self.refresh_list()

    def remove_selected(self):
        # 対面・遠隔のリストで選択があればそちらを削除
        if self.listbox_face.curselection():
            self.face_file = None
            self.refresh_listbox_alkira()
            return
        if self.listbox_remote.curselection():
            self.remote_file = None
            self.refresh_listbox_alkira()
            return
        # 運行データのリストで選択があれば削除
        sel = list(self.listbox.curselection())
        if not sel:
            return
        for idx in reversed(sel):
            del self.files[idx]
        self.refresh_list()

    def clear_all(self):
        self.files = []
        self.face_file = None
        self.remote_file = None
        self.refresh_list()
        self.refresh_listbox_alkira()

    def select_face_file(self):
        """対面アルキラーファイル"""
        path = filedialog.askopenfilename(
            title="対面ファイル",
            filetypes=[
                ("Excel / CSV files", "*.xlsx *.xlsm *.csv"),
                ("Excel files", "*.xlsx *.xlsm"),
                ("CSV files", "*.csv"),
                ("All files", "*.*"),
            ],
        )
        if path:
            self.face_file = path
            self.refresh_listbox_alkira()

    def select_remote_file(self):
        """遠隔アルキラーファイル"""
        path = filedialog.askopenfilename(
            title="遠隔ファイル",
            filetypes=[
                ("Excel / CSV files", "*.xlsx *.xlsm *.csv"),
                ("Excel files", "*.xlsx *.xlsm"),
                ("CSV files", "*.csv"),
                ("All files", "*.*"),
            ],
        )
        if path:
            self.remote_file = path
            self.refresh_listbox_alkira()

    def move_up(self):
        sel = list(self.listbox.curselection())
        if not sel or sel[0] == 0:
            return
        for i in sel:
            self.files[i-1], self.files[i] = self.files[i], self.files[i-1]
        self.refresh_list()
        for i in [x-1 for x in sel]:
            self.listbox.selection_set(i)

    def move_down(self):
        sel = list(self.listbox.curselection())
        if not sel or sel[-1] == len(self.files) - 1:
            return
        for i in reversed(sel):
            self.files[i+1], self.files[i] = self.files[i], self.files[i+1]
        self.refresh_list()
        for i in [x+1 for x in sel]:
            self.listbox.selection_set(i)

    # ---------- 月末用 ----------
    def refresh_list_monthly(self):
        self.listbox_monthly.delete(0, tk.END)
        for f in self.monthly_files:
            self.listbox_monthly.insert(tk.END, os.path.basename(f))
        self.listbox_work_detail.delete(0, tk.END)
        for f in self.work_detail_files:
            self.listbox_work_detail.insert(tk.END, os.path.basename(f))

    def add_monthly_files(self):
        paths = filedialog.askopenfilenames(
            title="毎日の完成ファイルを選択",
            filetypes=[
                ("Excel / CSV files", "*.xlsx *.xlsm *.csv"),
                ("Excel files", "*.xlsx *.xlsm"),
                ("CSV files", "*.csv"),
                ("All files", "*.*"),
            ],
        )
        if paths:
            for p in paths:
                if p not in self.monthly_files:
                    self.monthly_files.append(p)
            self.refresh_list_monthly()

    def remove_selected_monthly(self):
        sel = list(self.listbox_monthly.curselection())
        if not sel:
            return
        for idx in reversed(sel):
            del self.monthly_files[idx]
        self.refresh_list_monthly()

    def clear_all_monthly(self):
        self.monthly_files = []
        self.work_detail_files = []
        self.refresh_list_monthly()

    def add_work_detail_files(self):
        """作業明細ファイル（複数選択可・分割休息②用）"""
        paths = filedialog.askopenfilenames(
            title="作業明細ファイルを選択",
            filetypes=[
                ("Excel / CSV files", "*.xlsx *.xlsm *.csv"),
                ("Excel files", "*.xlsx *.xlsm"),
                ("CSV files", "*.csv"),
                ("All files", "*.*"),
            ],
        )
        if paths:
            for p in paths:
                if p not in self.work_detail_files:
                    self.work_detail_files.append(p)
            self.refresh_list_monthly()

    def remove_selected_work_detail(self):
        sel = list(self.listbox_work_detail.curselection())
        if not sel:
            return
        for idx in reversed(sel):
            del self.work_detail_files[idx]
        self.refresh_list_monthly()

    def clear_all_work_detail(self):
        self.work_detail_files = []
        self.refresh_list_monthly()

    def merge_monthly_and_export(self):
        if len(self.monthly_files) < 1:
            messagebox.showwarning(APP_TITLE, "毎日ファイルが選択されていません。")
            return

        # 作業明細が未選択の場合、確認する
        if not self.work_detail_files:
            if not messagebox.askyesno(APP_TITLE, "作業明細ファイルが追加されていませんが、このまま続けますか？"):
                return

        filename = simpledialog.askstring(
            APP_TITLE,
            "完成したファイルの名前を入力してください",
            initialvalue="月末ファイル"
        )
        if not filename:
            return

        filename = os.path.splitext(filename)[0]
        output_filename = f"{filename}.xlsx"
        out_path = os.path.join(self.output_dir_monthly, output_filename)

        if os.path.isfile(out_path):
            if not messagebox.askyesno(APP_TITLE, "既に同じ名前のファイルが「完成フォルダ」にあります。このまま上書きしますか？"):
                return

        try:
            # ① 毎日ファイルを統合
            all_rows = merge_monthly_files(self.monthly_files)

            if not all_rows:
                messagebox.showwarning(APP_TITLE, "統合するデータがありませんでした。")
                return

            self.monthly_merged_data = all_rows

            # ② 作業明細が複数なら統合（先頭のヘッダーのみ残す）
            work_detail_path = None
            temp_work_detail = None
            if self.work_detail_files:
                work_detail_path = _merge_work_detail_files(self.work_detail_files)
                if len(self.work_detail_files) > 1 and work_detail_path:
                    temp_work_detail = work_detail_path

            # ③ 分割休息①→②を実行
            final_rows = run_split_rest(all_rows, work_detail_path)

            if temp_work_detail and os.path.isfile(temp_work_detail):
                try:
                    os.remove(temp_work_detail)
                except Exception:
                    pass

            # ③ 完成フォルダに出力
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.title = OUTPUT_SHEETNAME
            for row in final_rows:
                ws_out.append(row)

            # 点呼日時列（U, W, Y）の表示形式
            from datetime import datetime as dt_cls
            date_fmt = "yyyy/m/d h:mm"
            for col in (21, 23, 25):
                for cell in ws_out.iter_cols(min_col=col, max_col=col, min_row=2):
                    for c in cell:
                        if isinstance(c.value, dt_cls):
                            c.number_format = date_fmt

            # 分割開始・終了列（Q, R, S, T）は yyyy/m/d h:mm:ss
            split_fmt = "yyyy/m/d h:mm:ss"
            for col in (17, 18, 19, 20):
                for c in ws_out.iter_cols(min_col=col, max_col=col, min_row=2):
                    for cell in c:
                        cell.number_format = split_fmt

            # 休憩・休息・休息採用列（J, K, L）は時刻表示 [h]:mm:ss（シリアル値を正しく表示）
            time_fmt = "[h]:mm:ss"
            for col in (10, 11, 12):
                for c in ws_out.iter_cols(min_col=col, max_col=col, min_row=2):
                    for cell in c:
                        cell.number_format = time_fmt

            yu_gothic_font = Font(name="游ゴシック")
            for row in ws_out.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.font = yu_gothic_font

            wb_out.save(out_path)
            data_rows = len(final_rows) - 1
            detail_msg = "作業明細あり" if self.work_detail_files else "作業明細なし（分割②はスキップ）"
            messagebox.showinfo(APP_TITLE, f"完了しました。\n\n保存先：「完成フォルダ」内の「月末」に保存しました。\n各ファイルは「ゴミ箱」に移動しました。")

            # 使用したファイルを「ゴミ箱」フォルダへ移動（毎日ファイル + 作業明細）
            all_src_paths = list(dict.fromkeys(
                list(self.monthly_files) + list(self.work_detail_files)
            ))
            for src_path in all_src_paths:
                if not os.path.isfile(src_path):
                    continue
                base_name = os.path.basename(src_path)
                dst_path = os.path.join(self.trash_dir, base_name)
                if os.path.exists(dst_path):
                    name, ext = os.path.splitext(base_name)
                    counter = 1
                    while True:
                        candidate = f"{name}_{counter}{ext}"
                        candidate_path = os.path.join(self.trash_dir, candidate)
                        if not os.path.exists(candidate_path):
                            dst_path = candidate_path
                            break
                        counter += 1
                try:
                    shutil.move(src_path, dst_path)
                except Exception:
                    pass
            self.monthly_files = []
            self.work_detail_files = []
            self.refresh_list_monthly()

        except Exception as e:
            messagebox.showerror(APP_TITLE, _user_friendly_error_message(e))

    def merge_and_export(self):
        if len(self.files) < 1:
            messagebox.showwarning(APP_TITLE, "運行データファイルが選択されていません。")
            return

        # 3種類がそろっていない場合、確認する（運行データは上でチェック済み）
        missing = []
        if not self.face_file:
            missing.append("対面アルキラー")
        if not self.remote_file:
            missing.append("遠隔アルキラー")
        if missing:
            msg = "「" + "」「".join(missing) + "」が追加されていませんが、このまま続けますか？"
            if not messagebox.askyesno(APP_TITLE, msg):
                return

        # ファイル名をユーザーに入力してもらう
        filename = simpledialog.askstring(
            APP_TITLE,
            "完成したファイルの名前を入力してください",
            initialvalue="毎日ファイル"
        )
        
        if not filename:
            return  # キャンセルされた場合
        
        # 拡張子が含まれていれば削除
        filename = os.path.splitext(filename)[0]
        output_filename = f"{filename}.xlsx"
        out_path = os.path.join(self.output_dir_daily, output_filename)

        if os.path.isfile(out_path):
            if not messagebox.askyesno(APP_TITLE, "既に同じ名前のファイルが「完成フォルダ」にあります。このまま上書きしますか？"):
                return

        try:
            all_rows = []
            first_file = True

            for path in self.files:
                _, ext = os.path.splitext(path)
                ext = ext.lower()

                # ==============================
                # Excel ファイル（.xlsx / .xlsm）
                # ==============================
                if ext in [".xlsx", ".xlsm"]:
                    wb_in = load_workbook(path, data_only=True)
                    ws_in = wb_in.worksheets[0]  # 先頭シート

                    max_row = ws_in.max_row
                    max_col = ws_in.max_column

                    if max_row is None or max_row == 0:
                        continue

                    start_row = 1 if first_file else 2  # 先頭以外はヘッダー(1行目)を飛ばす

                    for r in range(start_row, max_row + 1):
                        row_values = [ws_in.cell(row=r, column=c).value for c in range(1, max_col + 1)]

                        # 完全空行はスキップ
                        if all(v is None or v == "" for v in row_values):
                            continue

                        all_rows.append(row_values)

                    first_file = False

                # ==========
                # CSV ファイル
                # ==========
                elif ext == ".csv":
                    # 文字コードは日本語環境想定で cp932 を優先
                    # 読めない場合は utf-8 で再トライ
                    rows = []
                    try:
                        with open(path, "r", encoding="cp932", newline="") as f:
                            reader = csv.reader(f)
                            rows = list(reader)
                    except UnicodeDecodeError:
                        with open(path, "r", encoding="utf-8", newline="") as f:
                            reader = csv.reader(f)
                            rows = list(reader)

                    if not rows:
                        continue

                    start_idx = 0 if first_file else 1  # 先頭以外はヘッダー(1行目)を飛ばす

                    for r in rows[start_idx:]:
                        # 完全空行はスキップ
                        if not any(cell not in (None, "", " ") for cell in r):
                            continue

                        # 数値っぽい文字列は int / float に変換してから追加
                        converted_row = [_convert_cell_value(cell) for cell in r]
                        all_rows.append(converted_row)

                    first_file = False

                # それ以外の拡張子はスキップ
                else:
                    continue

            # 指定列のみに絞り込み
            filtered_rows = _filter_rows_by_columns(all_rows, KEEP_COLUMNS, INSERT_COLUMNS)

            # 車両コードの一の桁を枝番列に転記
            _fill_eda_num_from_vehicle_code(filtered_rows, KEEP_COLUMNS)

            # 点呼データ転記（対面・遠隔が両方選択されている場合）
            if self.face_file and self.remote_file:
                tenko_result = integrate_tenko_data(self.face_file, self.remote_file)
                fill_tenko_into_rows(
                    filtered_rows, tenko_result, KEEP_COLUMNS,
                    dep_minutes=self.tenko_dep_minutes, ret_minutes=self.tenko_ret_minutes
                )

            # 運転列（AA）を 5:40:47 形式に整形（先頭0なし・小数点以下なし）
            try:
                idx_untin = KEEP_COLUMNS.index("運転")
                for row in filtered_rows[1:]:
                    if len(row) > idx_untin:
                        row[idx_untin] = _format_untin_display(row[idx_untin])
            except ValueError:
                pass

            # 出力用ワークブックに書き込み
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.title = OUTPUT_SHEETNAME_DAILY
            for row in filtered_rows:
                ws_out.append(row)

            # 点呼日時列（U, W, Y）の表示形式
            from datetime import datetime as dt_cls
            date_fmt = "yyyy/m/d h:mm"
            for col in (21, 23, 25):  # U, W, Y列
                for cell in ws_out.iter_cols(min_col=col, max_col=col, min_row=2):
                    for c in cell:
                        if isinstance(c.value, dt_cls):
                            c.number_format = date_fmt

            # 分割開始・終了列（Q, R, S, T）は yyyy/m/d h:mm:ss
            split_fmt = "yyyy/m/d h:mm:ss"
            for col in (17, 18, 19, 20):
                for c in ws_out.iter_cols(min_col=col, max_col=col, min_row=2):
                    for cell in c:
                        cell.number_format = split_fmt

            # 休憩・休息・休息採用列（J, K, L）は時刻表示 [h]:mm:ss
            time_fmt = "[h]:mm:ss"
            for col in (10, 11, 12):
                for c in ws_out.iter_cols(min_col=col, max_col=col, min_row=2):
                    for cell in c:
                        cell.number_format = time_fmt

            # すべてのセルのフォントを游ゴシックに統一
            yu_gothic_font = Font(name="游ゴシック")
            for row in ws_out.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.font = yu_gothic_font

            wb_out.save(out_path)

            # 元ファイルを「ゴミ箱」フォルダへ移動（SGファイル + 対面・遠隔）
            all_src_paths = list(dict.fromkeys(
                list(self.files)
                + ([self.face_file] if self.face_file else [])
                + ([self.remote_file] if self.remote_file else [])
            ))
            for src_path in all_src_paths:
                if not os.path.isfile(src_path):
                    continue
                base_name = os.path.basename(src_path)
                dst_path = os.path.join(self.trash_dir, base_name)

                # 同名ファイルが既にある場合は _1, _2, ... を付与
                if os.path.exists(dst_path):
                    name, ext = os.path.splitext(base_name)
                    counter = 1
                    while True:
                        candidate = f"{name}_{counter}{ext}"
                        candidate_path = os.path.join(self.trash_dir, candidate)
                        if not os.path.exists(candidate_path):
                            dst_path = candidate_path
                            break
                        counter += 1

                try:
                    shutil.move(src_path, dst_path)
                except Exception:
                    # 移動に失敗しても処理は続行する
                    pass

            # ファイルリストをクリア
            self.files = []
            self.face_file = None
            self.remote_file = None
            self.refresh_list()

            messagebox.showinfo(APP_TITLE, f"完了しました。\n\n保存先：「完成フォルダ」内の「毎日」に保存しました。\n各ファイルは「ゴミ箱」に移動しました。")

        except Exception as e:
            messagebox.showerror(APP_TITLE, _user_friendly_error_message(e))

if __name__ == "__main__":
    App().mainloop()
