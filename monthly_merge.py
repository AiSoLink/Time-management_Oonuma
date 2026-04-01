"""
月末統合モジュール
毎日の完成ファイル（複数）を1つに統合する。
- 運行日が最早のファイルの乗務員名順を基準に統合
- 同じ乗務員名の行は運行日順に並び替え
"""

import csv
import os
from collections import defaultdict
from datetime import datetime, timedelta
from typing import List, Optional, Tuple

from openpyxl import load_workbook

# 完成ファイルの列インデックス（0-based）
COL_OPERATION_DATE = 1   # B列: 運行日
COL_CREW_NAME = 3       # D列: 乗務員名

EXCEL_EPOCH = datetime(1899, 12, 30)


def _parse_date(v) -> Optional[datetime]:
    """運行日を日付として解釈。比較用に datetime を返す。"""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    try:
        from datetime import date
        if isinstance(v, date) and not isinstance(v, datetime):
            return datetime.combine(v, datetime.min.time())
    except ImportError:
        pass
    if isinstance(v, str):
        v = v.strip()
        if not v:
            return None
        for fmt in ("%Y/%m/%d %H:%M:%S.%f", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M",
                    "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
                    "%Y/%m/%d", "%Y-%m-%d"):
            try:
                return datetime.strptime(v, fmt)
            except ValueError:
                continue
    try:
        n = float(v)
        if n > 0:
            return EXCEL_EPOCH + timedelta(days=n)
    except (TypeError, ValueError):
        pass
    return None


def _convert_cell_value(value) -> object:
    """
    CSV から読み込んだ文字列を、可能であれば数値(int / float)に変換する。
    """
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return ""

    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return value


def _read_file_to_rows(path: str) -> List[List]:
    """ファイル（Excel/CSV）を読み込み、行のリストを返す。"""
    _, ext = os.path.splitext(path)
    ext = ext.lower()

    if ext in [".xlsx", ".xlsm"]:
        wb = load_workbook(path, data_only=True)
        ws = wb.worksheets[0]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0:
            return []
        rows = []
        for r in range(1, max_row + 1):
            row = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            rows.append(row)
        return rows

    if ext == ".csv":
        try:
            with open(path, "r", encoding="cp932", newline="") as f:
                rows = list(csv.reader(f))
        except UnicodeDecodeError:
            with open(path, "r", encoding="utf-8", newline="") as f:
                rows = list(csv.reader(f))
        # 数値変換
        return [[_convert_cell_value(cell) for cell in row] for row in rows]

    return []


def _is_empty_row(row: List) -> bool:
    return not any(cell not in (None, "", " ") for cell in row)


def _get_crew_name(row: List) -> str:
    """行から乗務員名を取得"""
    if COL_CREW_NAME >= len(row):
        return ""
    v = row[COL_CREW_NAME]
    return str(v).strip() if v is not None else ""


def merge_monthly_files(paths: List[str]) -> List[List]:
    """
    毎日の完成ファイル（複数）を1つに統合する。

    - 運行日(B列)が最早のファイルを1番目に
    - 1番目のファイルの乗務員名(D列)順を基準に、同じ乗務員名の行は運行日順で並べる
    - 1番目に無い乗務員名は末尾に、2番目以降で初登場の順で追加

    Args:
        paths: ファイルパスのリスト

    Returns:
        [ヘッダー行, データ行1, データ行2, ...]
    """
    if not paths:
        return []

    # 全ファイルを読み込み
    file_rows_list: List[Tuple[List, List[List]]] = []
    for path in paths:
        rows = _read_file_to_rows(path)
        if not rows:
            continue
        header = rows[0]
        data_rows = [r for r in rows[1:] if not _is_empty_row(r)]
        if not data_rows:
            continue
        file_rows_list.append((header, data_rows))

    if not file_rows_list:
        return []

    # 各ファイルの最早運行日を算出し、1番目を決定
    def get_min_date(rows: List[List]) -> Optional[datetime]:
        dates = []
        for row in rows:
            if COL_OPERATION_DATE < len(row):
                dt = _parse_date(row[COL_OPERATION_DATE])
                if dt is not None:
                    dates.append(dt)
        return min(dates) if dates else None

    file_rows_list.sort(key=lambda x: get_min_date(x[1]) or datetime.max)

    header = file_rows_list[0][0]
    first_data = file_rows_list[0][1]
    rest_data = []
    for _, data in file_rows_list[1:]:
        rest_data.extend(data)

    # 1番目ファイルの乗務員名順（初登場順）
    crew_order = []
    seen = set()
    for row in first_data:
        name = _get_crew_name(row)
        if name and name not in seen:
            crew_order.append(name)
            seen.add(name)

    # 2番目以降にのみ登場する乗務員名（初登場順）
    for row in rest_data:
        name = _get_crew_name(row)
        if name and name not in seen:
            crew_order.append(name)
            seen.add(name)

    # 乗務員名ごとに全行を集約
    name_to_rows = defaultdict(list)
    for row in first_data + rest_data:
        name = _get_crew_name(row)
        if name:
            name_to_rows[name].append(row)

    # 乗務員名順 × 運行日順で出力
    result = [header]
    for crew_name in crew_order:
        rows = name_to_rows.get(crew_name, [])
        rows.sort(key=lambda r: (_parse_date(r[COL_OPERATION_DATE]) if COL_OPERATION_DATE < len(r) else datetime.min) or datetime.min)
        result.extend(rows)

    return result
